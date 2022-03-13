from os import path
import argparse
from datetime import datetime
from openpyxl import load_workbook

from table import Table

parser = argparse.ArgumentParser(prog='Michal', description='Process some compounds.')
parser.add_argument('-c', '--compound', nargs='+',
                    help='compound or compounds name')
parser.add_argument('--workbook',
                    help='xlsx workbook path')
parser.add_argument('-o', '--out',
                    help='xlsx workbook output path')

args = parser.parse_args()

if args.compound:
    for compound in args.compound:
        print(f'PubChem, {compound} => {Table.getInfoPubChem(compound)}')
        print(f'Nist, {compound} => {Table.getInfoNIST(compound)}')
elif args.workbook:
    output_path = args.out or path.join(path.dirname(args.workbook),f"""{datetime.now().strftime('%d%H%M%S')}.xlsx""")
    wb = load_workbook(filename = args.workbook)
    print(wb.sheetnames)
    ws = wb['מרוכז חדש']
    wb.save(output_path)
    # TODO
else:
    parser.print_usage()