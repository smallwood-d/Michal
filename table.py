from typing import Tuple
from openpyxl import worksheet
from collections import namedtuple
from pubchempy import get_compounds as pub_get_compounds

from nist import get_compounds as nist_get_compounds, Info


Cell = namedtuple('Cell', ['row', 'col', 'val'])

class Table:
    """Fill Michal xlsx tables"""
    def __init__(self, worksheet: worksheet, compound_cell: Cell) -> None:
        self.compound_cell = compound_cell
        self.formula = Cell(compound_cell.row, compound_cell.col - 1, "formula")
        self.Ki = Cell(compound_cell.row, compound_cell.col - 2, "Ki")
        self.active_phase = Cell(compound_cell.row, compound_cell.col - 3, "active phase")
        self.current_pos = 0
        self.worksheet = worksheet

    def fill_table(self) -> None:
        for i in range (self.compound_cell.row, self.worksheet.max_row):
            if i % 10 == 0:
                print(i)
            curr_value = self.worksheet.cell(i,self.compound_cell.col).value
            
            if curr_value:
                Table.getInfo(curr_value)
                new_name = Table.getInfoPubChem(curr_value)
                if new_name and new_name[0][0]:
                    self.worksheet.cell(i,self.formula.col).value = new_name[0][0]

    @classmethod
    def find_compound_cells(cls, worksheet: worksheet):
        for row in worksheet.rows:
            for cell in row:
                if cell.value == "Compound":
                    yield cls(worksheet, Cell(cell.row, cell.column, "Compound"))

    @staticmethod
    def getInfoPubChem(name: str) -> list[Tuple[str, str]]:
        """collect all the synonyms Names from PubChem."""
        realNames = []
        if name:
            for compound in pub_get_compounds(name, 'name'):
                realName = compound.synonyms
                if realName:
                    realNames.append((compound.molecular_formula, realName[0]))
        return realNames

    @staticmethod
    def getInfoNIST(name: str) -> Info:
            return nist_get_compounds(name)

    @staticmethod
    def getInfo(name: str) -> None:
        print(Table.getInfoNIST(name))
        print(Table.getInfoPubChem(name))